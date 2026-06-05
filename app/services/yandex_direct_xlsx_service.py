"""Parse Yandex Direct Master Report / campaign statistics .xlsx into an audit data slice."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from app.services.metrics_extract_service import parse_num

# Minimum header signals to treat sheet as a campaign statistics table.
_HEADER_SCORE_MIN = 4
_MAX_META_ROWS = 25
_MAX_SCAN_ROWS = 5000

CONDITIONS_MIN_COST_RUB = 500
CONDITIONS_MIN_LEADS_CPL = 3
CONDITIONS_TOP_SPEND = 20
CONDITIONS_TOP_WASTE = 15
CONDITIONS_TOP_CPL_BEST = 10
CONDITIONS_TOP_CPL_WORST = 10
CONDITIONS_TOP_MONTHLY = 5
CONDITIONS_TOP_PER_CAMPAIGN = 10
CONDITIONS_DISPLAY_MAX_LEN = 200

_PERIOD_RE = re.compile(
    r"\b(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})\s*[-—]\s*(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})\b"
)
_FILENAME_PERIOD_RE = re.compile(
    r"(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})"
)


def _norm_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _norm_header(value: Any) -> str:
    text = _norm_cell(value).lower()
    text = text.replace("₽", "").replace("%", "").strip()
    return text


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return parse_num(str(value))


def _parse_int(value: Any) -> int | None:
    num = _parse_float(value)
    if num is None:
        return None
    return int(round(num))


def _column_kind(header: str) -> str | None:
    h = header
    if not h:
        return None
    if h.startswith("№") or "номер кампании" in h or h in {"№", "id", "id кампании"}:
        return "campaign_id"
    if "название кампании" in h or h == "кампания" or h.startswith("кампания"):
        return "campaign_name"
    if "групп" in h and "кампан" not in h:
        return None
    if "расход" in h or "затрат" in h or h == "cost":
        return "cost"
    if h == "клики" or h.startswith("клик"):
        return "clicks"
    if "конверс" in h or "заявк" in h or "целев" in h:
        return "conversions"
    if h.startswith("cr") or "конверсия" in h:
        return "cr"
    if "cpa" in h or "cpl" in h or "цена конверс" in h:
        return "cpa"
    if "показ" in h or h == "impressions":
        return "impressions"
    if "ctr" in h:
        return "ctr"
    if "cpc" in h or "ср. цена клика" in h:
        return "cpc"
    return None


def _map_columns(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        h = _norm_header(raw)
        if h == "месяц" or h.startswith("месяц"):
            mapping["month"] = idx
        kind = _column_kind(h)
        if kind and kind not in mapping:
            mapping[kind] = idx
    return mapping


def _header_score(headers: list[Any]) -> int:
    kinds = {_column_kind(_norm_header(h)) for h in headers}
    kinds.discard(None)
    score = 0
    if "cost" in kinds:
        score += 2
    if "clicks" in kinds:
        score += 2
    if "conversions" in kinds:
        score += 2
    if "campaign_name" in kinds or "campaign_id" in kinds:
        score += 1
    if "cpa" in kinds or "cr" in kinds:
        score += 1
    return score


def _row_value(row: tuple[Any, ...], col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_total_row(row: tuple[Any, ...]) -> bool:
    for cell in row[:3]:
        text = _norm_cell(cell).lower()
        if text and "итого" in text:
            return True
    return False


def _scan_meta(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "client_label": None,
        "report_title": None,
        "period": None,
    }
    for row in rows[:_MAX_META_ROWS]:
        line = " | ".join(_norm_cell(c) for c in row if _norm_cell(c))
        if not line:
            continue
        low = line.lower()
        if meta["client_label"] is None and re.search(r"\bклиент\s*:", low):
            meta["client_label"] = re.sub(r"(?i)^клиент\s*:\s*", "", line).strip()
        if meta["report_title"] is None and ("отчет" in low or "отчёт" in low or "мастер" in low):
            meta["report_title"] = line
        if meta["period"] is None:
            m = _PERIOD_RE.search(line)
            if m:
                meta["period"] = f"{m.group(1)} — {m.group(2)}".replace("-", "—")
    return meta


def _period_from_filename(path: Path) -> str | None:
    m = _FILENAME_PERIOD_RE.search(path.name)
    if not m:
        return None
    return f"{m.group(1)} — {m.group(2)}"


_MONTH_SORT = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]


def _month_sort_key(label: str) -> tuple[int, int]:
    text = (label or "").lower()
    for i, prefix in enumerate(_MONTH_SORT):
        if text.startswith(prefix):
            try:
                year = int(text.split()[-1])
            except (ValueError, IndexError):
                year = 0
            return (year, i)
    return (0, 99)


def _is_conditions_report(headers: list[Any]) -> bool:
    joined = " ".join(_norm_header(h) for h in headers)
    return "условие показа" in joined and (
        ("конверсии" in joined and "мессендж" in joined)
        or ("конверсии" in joined and "форм" in joined)
    )


def _map_conditions_columns(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        h = _norm_header(raw)
        if not h:
            continue
        if h == "дата" or h.startswith("дата"):
            mapping["month"] = idx
        elif h == "кампания":
            mapping["campaign_name"] = idx
        elif "№" in h and "кампани" in h:
            mapping["campaign_id"] = idx
        elif "условие показа" in h:
            mapping["condition"] = idx
        elif h == "показы" or h.startswith("показ"):
            mapping["impressions"] = idx
        elif h == "клики" or h.startswith("клик"):
            mapping["clicks"] = idx
        elif "расход" in h:
            mapping["cost"] = idx
        elif "конверсии" in h and "форм" in h:
            mapping["conv_forms"] = idx
        elif "конверсии" in h and "мессендж" in h:
            mapping["conv_messenger"] = idx
        elif "цена цели" in h and "форм" in h:
            mapping["cpa_form"] = idx
        elif "цена цели" in h and "мессендж" in h:
            mapping["cpa_messenger"] = idx
    return mapping


def _classify_condition_kind(condition_text: str) -> str:
    text = (condition_text or "").strip().lower()
    if not text or text in {"—", "-", "–"}:
        return "empty"
    if "автотаргет" in text:
        return "autotarget"
    return "keyword"


def _normalize_condition_key(condition_raw: str) -> tuple[str, str]:
    display = _norm_cell(condition_raw) or "—"
    key = display.lower()
    return key, display[:CONDITIONS_DISPLAY_MAX_LEN]


def _new_condition_bucket() -> dict[str, Any]:
    return {
        "display": "",
        "cost": 0.0,
        "clicks": 0,
        "impressions": 0,
        "forms": 0,
        "messengers": 0,
        "top_campaign_id": "",
        "top_campaign_name": "",
        "top_campaign_cost": 0.0,
    }


def _accumulate_condition_bucket(
    bucket: dict[str, Any],
    *,
    cost: float,
    clicks: int,
    impressions: int,
    forms: int,
    messengers: int,
    display: str | None = None,
    campaign_id: str = "",
    campaign_name: str = "",
) -> None:
    if display and not bucket.get("display"):
        bucket["display"] = display
    bucket["cost"] += cost
    bucket["clicks"] += clicks
    bucket["impressions"] += impressions
    bucket["forms"] += forms
    bucket["messengers"] += messengers
    if campaign_id and cost >= bucket.get("top_campaign_cost", 0):
        bucket["top_campaign_cost"] = cost
        bucket["top_campaign_id"] = campaign_id
        bucket["top_campaign_name"] = campaign_name or bucket.get("top_campaign_name") or ""


def _campaign_condition_totals(campaign_row: dict[str, Any]) -> dict[str, Any]:
    leads = int(campaign_row.get("leads") or 0)
    cost = float(campaign_row.get("cost") or 0)
    totals: dict[str, Any] = {"cost": cost, "leads": leads}
    if leads > 0:
        totals["cpl"] = round(cost / leads, 2)
    return totals


def _build_conditions_summary(
    by_condition: dict[str, dict[str, Any]],
    *,
    rows_in_file: int,
    totals: dict[str, Any],
    top_spend_limit: int = CONDITIONS_TOP_SPEND,
    top_waste_limit: int = CONDITIONS_TOP_WASTE,
) -> dict[str, Any]:
    total_cost = float(totals.get("cost") or 0)
    total_leads = int(totals.get("leads") or 0)

    rows_out: list[dict[str, Any]] = []
    for key, d in by_condition.items():
        leads = int(d.get("forms") or 0) + int(d.get("messengers") or 0)
        cost = round(float(d.get("cost") or 0), 2)
        if cost <= 0 and leads <= 0:
            continue
        display = (d.get("display") or key)[:CONDITIONS_DISPLAY_MAX_LEN]
        rows_out.append(
            {
                "condition": display,
                "condition_kind": _classify_condition_kind(display),
                "campaign_id": d.get("top_campaign_id") or None,
                "campaign_name": d.get("top_campaign_name") or "",
                "cost": cost,
                "clicks": int(d.get("clicks") or 0),
                "impressions": int(d.get("impressions") or 0),
                "forms": int(d.get("forms") or 0),
                "messengers": int(d.get("messengers") or 0),
                "leads": leads,
                "cpl": round(cost / leads, 2) if leads > 0 else None,
                "share_of_cost_pct": round(100 * cost / total_cost, 2) if total_cost > 0 else None,
                "share_of_leads_pct": round(100 * leads / total_leads, 2) if total_leads > 0 else None,
            }
        )

    by_kind: dict[str, dict[str, Any]] = {
        "keyword": {"cost": 0.0, "leads": 0},
        "autotarget": {"cost": 0.0, "leads": 0},
        "other": {"cost": 0.0, "leads": 0},
    }
    for row in rows_out:
        kind = row.get("condition_kind") or "other"
        bucket = by_kind["autotarget"] if kind == "autotarget" else (
            by_kind["other"] if kind in {"empty", "other"} else by_kind["keyword"]
        )
        bucket["cost"] += row.get("cost") or 0
        bucket["leads"] += row.get("leads") or 0

    for bucket in by_kind.values():
        bucket["cost"] = round(bucket["cost"], 2)
        bucket["share_cost_pct"] = round(100 * bucket["cost"] / total_cost, 1) if total_cost > 0 else 0

    top_by_spend = sorted(rows_out, key=lambda r: -(r.get("cost") or 0))[:top_spend_limit]
    high_spend_zero_leads = sorted(
        [r for r in rows_out if (r.get("cost") or 0) >= CONDITIONS_MIN_COST_RUB and (r.get("leads") or 0) == 0],
        key=lambda r: -(r.get("cost") or 0),
    )[:top_waste_limit]
    cpl_ranked = [
        r for r in rows_out
        if (r.get("leads") or 0) >= CONDITIONS_MIN_LEADS_CPL and r.get("cpl") is not None
    ]
    top_best_cpl = sorted(cpl_ranked, key=lambda r: r.get("cpl") or 0)[:CONDITIONS_TOP_CPL_BEST]
    top_worst_cpl = sorted(cpl_ranked, key=lambda r: -(r.get("cpl") or 0))[:CONDITIONS_TOP_CPL_WORST]

    return {
        "rows_in_file": rows_in_file,
        "unique_conditions": len(by_condition),
        "totals": {
            "cost": total_cost,
            "leads": total_leads,
            "cpl": totals.get("cpl"),
        },
        "top_by_spend": top_by_spend,
        "high_spend_zero_leads": high_spend_zero_leads,
        "top_best_cpl": top_best_cpl,
        "top_worst_cpl": top_worst_cpl,
        "by_condition_kind": by_kind,
        "limits": {
            "min_cost_rub": CONDITIONS_MIN_COST_RUB,
            "min_leads_for_cpl_rank": CONDITIONS_MIN_LEADS_CPL,
        },
    }


def build_conditions_insights(conditions_summary: dict[str, Any]) -> list[dict[str, str]]:
    insights: list[dict[str, str]] = []
    if not conditions_summary:
        return insights
    top = conditions_summary.get("top_by_spend") or []
    waste = conditions_summary.get("high_spend_zero_leads") or []
    totals = conditions_summary.get("totals") or {}
    by_kind = conditions_summary.get("by_condition_kind") or {}

    if len(top) >= 3:
        top3_cost = sum(r.get("cost") or 0 for r in top[:3])
        total_cost = float(totals.get("cost") or 0)
        if total_cost > 0:
            pct = round(100 * top3_cost / total_cost, 1)
            insights.append(
                {
                    "id": "top3_conditions_share",
                    "title": "Концентрация бюджета",
                    "detail": f"Топ-3 условия показа = {pct}% расхода за период файла.",
                }
            )

    auto = by_kind.get("autotarget") or {}
    if auto.get("cost") and totals.get("cost"):
        insights.append(
            {
                "id": "autotarget_share",
                "title": "Автотаргетинг",
                "detail": (
                    f"Доля расхода на автотаргетинг: {auto.get('share_cost_pct', 0)}% "
                    f"({ _fmt_money(auto.get('cost')) }, {auto.get('leads', 0)} лидов)."
                ),
            }
        )

    if waste:
        waste_sum = sum(r.get("cost") or 0 for r in waste)
        insights.append(
            {
                "id": "conditions_zero_leads",
                "title": "Расход без лидов",
                "detail": (
                    f"{len(waste)} условий с расходом ≥ {CONDITIONS_MIN_COST_RUB} ₽ и 0 лидов "
                    f"(сумма {_fmt_money(waste_sum)}) — кандидаты на минус/паузу."
                ),
            }
        )

    best = (conditions_summary.get("top_best_cpl") or [])[:1]
    if best:
        row = best[0]
        insights.append(
            {
                "id": "best_condition_cpl",
                "title": "Лучшее условие по CPL",
                "detail": (
                    f"«{(row.get('condition') or '')[:60]}» — {_fmt_money(row.get('cpl'))} "
                    f"({row.get('leads')} лидов), кампания {row.get('campaign_name') or '—'}."
                ),
            }
        )
    return insights


def slim_conditions_for_ai(conditions_summary: dict[str, Any] | None) -> dict[str, Any]:
    """Subset for AI prompt — no full unique_conditions list."""
    if not conditions_summary:
        return {}
    return {
        "unique_conditions": conditions_summary.get("unique_conditions"),
        "totals": conditions_summary.get("totals"),
        "top_by_spend": conditions_summary.get("top_by_spend"),
        "high_spend_zero_leads": conditions_summary.get("high_spend_zero_leads"),
        "top_best_cpl": conditions_summary.get("top_best_cpl"),
        "top_worst_cpl": conditions_summary.get("top_worst_cpl"),
        "by_condition_kind": conditions_summary.get("by_condition_kind"),
        "insights": conditions_summary.get("insights"),
        "monthly_tops": (conditions_summary.get("monthly_tops") or [])[:6],
        "by_campaign": (conditions_summary.get("by_campaign") or [])[:12],
    }


def _build_conditions_by_campaign(
    by_campaign_condition: dict[str, dict[str, dict[str, Any]]],
    campaigns: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    campaign_index = {
        str(c.get("campaign_id") if c.get("campaign_id") is not None else "unknown"): c
        for c in campaigns
    }
    out: list[dict[str, Any]] = []
    for cid, ckey_map in sorted(
        by_campaign_condition.items(),
        key=lambda item: -sum(b.get("cost") or 0 for b in item[1].values()),
    ):
        camp = campaign_index.get(cid) or {}
        camp_name = camp.get("campaign_name") or cid
        if camp:
            share_totals = _campaign_condition_totals(camp)
        else:
            cost_sum = sum(float(b.get("cost") or 0) for b in ckey_map.values())
            leads_sum = sum(int(b.get("forms") or 0) + int(b.get("messengers") or 0) for b in ckey_map.values())
            share_totals = {"cost": round(cost_sum, 2), "leads": leads_sum}
            if leads_sum > 0:
                share_totals["cpl"] = round(cost_sum / leads_sum, 2)
        sub = _build_conditions_summary(
            ckey_map,
            rows_in_file=0,
            totals=share_totals,
            top_spend_limit=CONDITIONS_TOP_PER_CAMPAIGN,
            top_waste_limit=min(8, CONDITIONS_TOP_WASTE),
        )
        if not sub.get("top_by_spend"):
            continue
        out.append(
            {
                "campaign_id": camp.get("campaign_id") if camp else (None if cid == "unknown" else cid),
                "campaign_name": camp_name,
                "top_by_spend": sub.get("top_by_spend") or [],
                "high_spend_zero_leads": sub.get("high_spend_zero_leads") or [],
                "top_best_cpl": sub.get("top_best_cpl") or [],
                "top_worst_cpl": sub.get("top_worst_cpl") or [],
            }
        )
    return out


def _build_monthly_condition_tops(
    by_month_condition: dict[str, dict[str, dict[str, Any]]],
    monthly: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    month_index = {str(m.get("month") or ""): m for m in monthly}
    tops: list[dict[str, Any]] = []
    for month in sorted(by_month_condition.keys(), key=_month_sort_key):
        ckey_map = by_month_condition[month]
        mrow = month_index.get(month) or {}
        leads = int(mrow.get("leads") or 0)
        cost = float(mrow.get("cost") or 0)
        totals = {"cost": cost, "leads": leads}
        if leads > 0:
            totals["cpl"] = mrow.get("cpl")
        sub = _build_conditions_summary(
            ckey_map,
            rows_in_file=0,
            totals=totals,
            top_spend_limit=CONDITIONS_TOP_MONTHLY,
            top_waste_limit=5,
        )
        if not sub.get("top_by_spend"):
            continue
        tops.append(
            {
                "month": month,
                "top_by_spend": sub.get("top_by_spend") or [],
                "high_spend_zero_leads": sub.get("high_spend_zero_leads") or [],
            }
        )
    return tops


def _parse_conditions_sheet(
    rows: list[tuple[Any, ...]],
    *,
    header_row: int,
    col_map: dict[str, int],
) -> dict[str, Any] | None:
    from collections import defaultdict

    if "cost" not in col_map:
        return None

    by_month: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"cost": 0.0, "clicks": 0, "impressions": 0, "forms": 0, "messengers": 0}
    )
    by_campaign: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"cost": 0.0, "clicks": 0, "impressions": 0, "forms": 0, "messengers": 0, "campaign_name": ""}
    )
    by_condition: dict[str, dict[str, Any]] = defaultdict(_new_condition_bucket)
    by_campaign_condition: dict[str, dict[str, dict[str, Any]]] = defaultdict(
        lambda: defaultdict(_new_condition_bucket)
    )
    by_month_condition: dict[str, dict[str, dict[str, Any]]] = defaultdict(
        lambda: defaultdict(_new_condition_bucket)
    )
    rows_in_file = 0
    has_condition_col = "condition" in col_map

    for row in rows[header_row + 1 :]:
        if not any(_norm_cell(c) for c in row):
            continue
        month = _norm_cell(_row_value(row, col_map, "month"))
        if not month:
            continue
        rows_in_file += 1
        cid = _norm_cell(_row_value(row, col_map, "campaign_id")) or "unknown"
        cname = _norm_cell(_row_value(row, col_map, "campaign_name"))
        cost = _parse_float(_row_value(row, col_map, "cost")) or 0
        clicks = _parse_int(_row_value(row, col_map, "clicks")) or 0
        impressions = _parse_int(_row_value(row, col_map, "impressions")) or 0
        forms = _parse_int(_row_value(row, col_map, "conv_forms")) or 0
        messengers = _parse_int(_row_value(row, col_map, "conv_messenger")) or 0

        if has_condition_col:
            cond_raw = _norm_cell(_row_value(row, col_map, "condition"))
            ckey, cdisplay = _normalize_condition_key(cond_raw)
            _accumulate_condition_bucket(
                by_condition[ckey],
                cost=cost,
                clicks=clicks,
                impressions=impressions,
                forms=forms,
                messengers=messengers,
                display=cdisplay,
                campaign_id=cid if cid != "unknown" else "",
                campaign_name=cname,
            )
            _accumulate_condition_bucket(
                by_campaign_condition[cid][ckey],
                cost=cost,
                clicks=clicks,
                impressions=impressions,
                forms=forms,
                messengers=messengers,
                display=cdisplay,
                campaign_id=cid,
                campaign_name=cname,
            )
            _accumulate_condition_bucket(
                by_month_condition[month][ckey],
                cost=cost,
                clicks=clicks,
                impressions=impressions,
                forms=forms,
                messengers=messengers,
                display=cdisplay,
                campaign_id=cid,
                campaign_name=cname,
            )

        by_month[month]["cost"] += cost
        by_month[month]["clicks"] += clicks
        by_month[month]["impressions"] += impressions
        by_month[month]["forms"] += forms
        by_month[month]["messengers"] += messengers

        by_campaign[cid]["cost"] += cost
        by_campaign[cid]["clicks"] += clicks
        by_campaign[cid]["impressions"] += impressions
        by_campaign[cid]["forms"] += forms
        by_campaign[cid]["messengers"] += messengers
        if cname:
            by_campaign[cid]["campaign_name"] = cname

    if not by_month:
        return None

    monthly: list[dict[str, Any]] = []
    for month in sorted(by_month.keys(), key=_month_sort_key):
        d = by_month[month]
        leads = d["forms"] + d["messengers"]
        cost = round(d["cost"], 2)
        monthly.append(
            {
                "month": month,
                "cost": cost,
                "clicks": d["clicks"],
                "impressions": d["impressions"],
                "forms": d["forms"],
                "messengers": d["messengers"],
                "leads": leads,
                "cpl": round(cost / leads, 2) if leads > 0 else None,
            }
        )

    campaigns: list[dict[str, Any]] = []
    for cid, d in sorted(by_campaign.items(), key=lambda x: -x[1]["cost"]):
        leads = d["forms"] + d["messengers"]
        cost = round(d["cost"], 2)
        if cost <= 0 and leads <= 0:
            continue
        campaigns.append(
            {
                "campaign_id": None if cid == "unknown" else cid,
                "campaign_name": d["campaign_name"] or cid,
                "cost": cost,
                "clicks": d["clicks"],
                "impressions": d["impressions"],
                "forms": d["forms"],
                "messengers": d["messengers"],
                "leads": leads,
                "cpl": round(cost / leads, 2) if leads > 0 else None,
            }
        )

    totals = {
        "campaign_name": "Итого",
        "cost": round(sum(m["cost"] for m in monthly), 2),
        "clicks": sum(m["clicks"] for m in monthly),
        "impressions": sum(m["impressions"] for m in monthly),
        "forms": sum(m["forms"] for m in monthly),
        "messengers": sum(m["messengers"] for m in monthly),
        "leads": sum(m["leads"] for m in monthly),
    }
    if totals["leads"] > 0:
        totals["cpl"] = round(totals["cost"] / totals["leads"], 2)

    result: dict[str, Any] = {
        "report_type": "conditions_by_month",
        "monthly": monthly,
        "campaigns": campaigns,
        "totals": totals,
        "goals": {
            "forms": "Отправка формы",
            "messengers": "Переход в мессенджеры",
        },
    }
    if has_condition_col and by_condition:
        summary = _build_conditions_summary(
            by_condition,
            rows_in_file=rows_in_file,
            totals=totals,
        )
        summary["insights"] = build_conditions_insights(summary)
        summary["by_campaign"] = _build_conditions_by_campaign(by_campaign_condition, campaigns)
        summary["monthly_tops"] = _build_monthly_condition_tops(by_month_condition, monthly)
        result["conditions_summary"] = summary
    return result


def _monthly_rows_from_buckets(
    by_month: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    monthly: list[dict[str, Any]] = []
    for month in sorted(by_month.keys(), key=_month_sort_key):
        d = by_month[month]
        leads = int(d["forms"]) + int(d["messengers"])
        cost = round(float(d["cost"]), 2)
        monthly.append(
            {
                "month": month,
                "cost": cost,
                "clicks": int(d["clicks"]),
                "impressions": int(d["impressions"]),
                "forms": int(d["forms"]),
                "messengers": int(d["messengers"]),
                "leads": leads,
                "cpl": round(cost / leads, 2) if leads > 0 else None,
            }
        )
    return monthly


def _campaign_rows_from_buckets(
    by_campaign: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    campaigns: list[dict[str, Any]] = []
    for cid, d in sorted(by_campaign.items(), key=lambda x: -(x[1]["cost"])):
        leads = int(d["forms"]) + int(d["messengers"])
        cost = round(float(d["cost"]), 2)
        if cost <= 0 and leads <= 0 and int(d["clicks"]) <= 0:
            continue
        campaigns.append(
            {
                "campaign_id": None if cid == "unknown" else cid,
                "campaign_name": d["campaign_name"] or cid,
                "cost": cost,
                "clicks": int(d["clicks"]),
                "impressions": int(d["impressions"]),
                "forms": int(d["forms"]),
                "messengers": int(d["messengers"]),
                "leads": leads,
                "conversions": leads,
                "cpl": round(cost / leads, 2) if leads > 0 else None,
            }
        )
    return campaigns


def _parse_sheet_table(
    rows: list[tuple[Any, ...]],
) -> dict[str, Any] | None:
    from collections import defaultdict

    header_idx = None
    col_map: dict[str, int] = {}
    best_score = 0
    for i, row in enumerate(rows[:40]):
        score = _header_score(row)
        if score > best_score:
            best_score = score
            header_idx = i
            col_map = _map_columns(row)
    if header_idx is None or best_score < _HEADER_SCORE_MIN:
        return None
    if "cost" not in col_map and "clicks" not in col_map:
        return None

    if "month" in col_map:
        by_month: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"cost": 0.0, "clicks": 0, "impressions": 0, "forms": 0, "messengers": 0}
        )
        by_campaign: dict[str, dict[str, Any]] = defaultdict(
            lambda: {
                "cost": 0.0,
                "clicks": 0,
                "impressions": 0,
                "forms": 0,
                "messengers": 0,
                "campaign_name": "",
                "campaign_id": None,
            }
        )
        totals: dict[str, Any] | None = None
        for row in rows[header_idx + 1 : header_idx + 1 + _MAX_SCAN_ROWS]:
            if not any(_norm_cell(c) for c in row):
                continue
            if _is_total_row(row):
                totals = {
                    "campaign_name": "Итого",
                    "cost": _parse_float(_row_value(row, col_map, "cost")),
                    "clicks": _parse_int(_row_value(row, col_map, "clicks")),
                    "conversions": _parse_int(_row_value(row, col_map, "conversions")),
                    "cr": _parse_float(_row_value(row, col_map, "cr")),
                    "cpa": _parse_float(_row_value(row, col_map, "cpa")),
                    "impressions": _parse_int(_row_value(row, col_map, "impressions")),
                }
                conv = totals.get("conversions") or 0
                totals["leads"] = conv
                totals["forms"] = 0
                totals["messengers"] = 0
                continue
            month = _norm_cell(_row_value(row, col_map, "month"))
            if not month:
                continue
            cost = _parse_float(_row_value(row, col_map, "cost")) or 0.0
            clicks = _parse_int(_row_value(row, col_map, "clicks")) or 0
            impressions = _parse_int(_row_value(row, col_map, "impressions")) or 0
            conv = _parse_int(_row_value(row, col_map, "conversions")) or 0
            by_month[month]["cost"] += cost
            by_month[month]["clicks"] += clicks
            by_month[month]["impressions"] += impressions
            by_month[month]["forms"] += conv
            cid = _norm_cell(_row_value(row, col_map, "campaign_id")) or "unknown"
            cname = _norm_cell(_row_value(row, col_map, "campaign_name"))
            by_campaign[cid]["cost"] += cost
            by_campaign[cid]["clicks"] += clicks
            by_campaign[cid]["impressions"] += impressions
            by_campaign[cid]["forms"] += conv
            if cname:
                by_campaign[cid]["campaign_name"] = cname
            by_campaign[cid]["campaign_id"] = cid if cid != "unknown" else None

        monthly = _monthly_rows_from_buckets(by_month)
        campaigns = _campaign_rows_from_buckets(by_campaign)
        if not monthly and not campaigns and not totals:
            return None
        if totals is None and monthly:
            totals = {
                "campaign_name": "Итого (расчёт)",
                "cost": round(sum(m["cost"] for m in monthly), 2),
                "clicks": sum(m["clicks"] for m in monthly),
                "impressions": sum(m["impressions"] for m in monthly),
                "forms": sum(m["forms"] for m in monthly),
                "messengers": sum(m["messengers"] for m in monthly),
                "leads": sum(m["leads"] for m in monthly),
            }
            if totals["leads"] > 0:
                totals["cpl"] = round(totals["cost"] / totals["leads"], 2)
            totals["conversions"] = totals["leads"]
        return {
            "header_row": header_idx + 1,
            "columns": list(col_map.keys()),
            "monthly": monthly,
            "campaigns": campaigns,
            "totals": totals,
        }

    campaigns: list[dict[str, Any]] = []
    totals: dict[str, Any] | None = None

    for row in rows[header_idx + 1 : header_idx + 1 + _MAX_SCAN_ROWS]:
        if not any(_norm_cell(c) for c in row):
            continue
        if _is_total_row(row):
            totals = {
                "campaign_name": "Итого",
                "cost": _parse_float(_row_value(row, col_map, "cost")),
                "clicks": _parse_int(_row_value(row, col_map, "clicks")),
                "conversions": _parse_int(_row_value(row, col_map, "conversions")),
                "cr": _parse_float(_row_value(row, col_map, "cr")),
                "cpa": _parse_float(_row_value(row, col_map, "cpa")),
                "impressions": _parse_int(_row_value(row, col_map, "impressions")),
            }
            continue

        name = _norm_cell(_row_value(row, col_map, "campaign_name"))
        cid = _norm_cell(_row_value(row, col_map, "campaign_id"))
        cost = _parse_float(_row_value(row, col_map, "cost"))
        clicks = _parse_int(_row_value(row, col_map, "clicks"))
        if not name and not cid and cost is None and clicks is None:
            continue
        if not name and not cid:
            continue
        # Skip subtotal / section headers without metrics
        if cost is None and clicks is None:
            low_name = (name or cid).lower()
            if any(x in low_name for x in ("итого", "всего", "сводка")):
                continue

        conv = _parse_int(_row_value(row, col_map, "conversions")) or 0
        campaigns.append(
            {
                "campaign_id": cid or None,
                "campaign_name": name or cid,
                "cost": cost,
                "clicks": clicks,
                "conversions": conv,
                "leads": conv,
                "forms": conv,
                "messengers": 0,
                "cr": _parse_float(_row_value(row, col_map, "cr")),
                "cpa": _parse_float(_row_value(row, col_map, "cpa")),
                "impressions": _parse_int(_row_value(row, col_map, "impressions")),
            }
        )

    if not campaigns and not totals:
        return None

    if totals is None and campaigns:
        # Sum as fallback only if single obvious total row missing
        totals = {
            "campaign_name": "Итого (расчёт)",
            "cost": sum(c.get("cost") or 0 for c in campaigns),
            "clicks": sum(c.get("clicks") or 0 for c in campaigns),
            "conversions": sum(c.get("conversions") or 0 for c in campaigns),
            "cpa": None,
            "cr": None,
            "impressions": sum(c.get("impressions") or 0 for c in campaigns),
        }
        if totals["cost"] == 0 and totals["clicks"] == 0:
            totals = None

    return {
        "header_row": header_idx + 1,
        "columns": list(col_map.keys()),
        "campaigns": campaigns,
        "totals": totals,
    }


def _fmt_money(value: float | None) -> str:
    if value is None:
        return "—"
    text = f"{value:,.2f}".replace(",", " ").replace(".", ",")
    return f"{text} ₽"


def _fmt_int(value: int | float | None) -> str:
    if value is None:
        return "—"
    return f"{int(round(value)):,}".replace(",", " ")


def _fmt_pct(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}%".replace(".", ",")


def build_slice_insights(document_slice: dict[str, Any]) -> list[dict[str, str]]:
    """Rule-based patterns for UI and AI (not a substitute for full analysis)."""
    insights: list[dict[str, str]] = []
    monthly = document_slice.get("monthly") or []
    campaigns = document_slice.get("campaigns") or []
    if len(monthly) >= 2:
        by_leads = max(monthly, key=lambda m: m.get("leads") or 0)
        by_cpl = min(
            (m for m in monthly if m.get("cpl") is not None),
            key=lambda m: m.get("cpl") or 0,
            default=None,
        )
        worst_cpl = max(
            (m for m in monthly if m.get("cpl") is not None),
            key=lambda m: m.get("cpl") or 0,
            default=None,
        )
        if by_leads:
            insights.append(
                {
                    "id": "peak_leads_month",
                    "title": "Пик лидов",
                    "detail": f"{by_leads.get('month')}: {by_leads.get('leads')} лидов (форма {by_leads.get('forms')}, мессенджер {by_leads.get('messengers')}).",
                }
            )
        if by_cpl and worst_cpl and by_cpl.get("month") != worst_cpl.get("month"):
            insights.append(
                {
                    "id": "cpl_range",
                    "title": "Разброс CPL по месяцам",
                    "detail": f"Лучший: {by_cpl.get('month')} — {_fmt_money(by_cpl.get('cpl'))}; худший: {worst_cpl.get('month')} — {_fmt_money(worst_cpl.get('cpl'))}.",
                }
            )
        first, last = monthly[0], monthly[-1]
        if first.get("leads") and last.get("leads"):
            delta = (last.get("leads") or 0) - (first.get("leads") or 0)
            insights.append(
                {
                    "id": "leads_trend",
                    "title": "Тренд лидов",
                    "detail": f"От {first.get('month')} к {last.get('month')}: {delta:+d} лидов.",
                }
            )
    if campaigns:
        ranked = sorted(campaigns, key=lambda c: (c.get("cpl") is None, c.get("cpl") or 1e9))
        efficient = [c for c in ranked if c.get("leads") and c.get("cpl")]
        if efficient:
            best = efficient[0]
            insights.append(
                {
                    "id": "best_campaign_cpl",
                    "title": "Лучший CPL среди кампаний",
                    "detail": f"{best.get('campaign_name')}: {_fmt_money(best.get('cpl'))} при {best.get('leads')} лидах.",
                }
            )
        costly = sorted(
            [c for c in campaigns if c.get("leads") and c.get("cpl")],
            key=lambda c: -(c.get("cpl") or 0),
        )
        if costly:
            worst = costly[0]
            insights.append(
                {
                    "id": "worst_campaign_cpl",
                    "title": "Дорогой лид",
                    "detail": f"{worst.get('campaign_name')}: {_fmt_money(worst.get('cpl'))} — кандидат на оптимизацию бюджета или семантики.",
                }
            )
    return insights


def slice_monthly_to_metrics_list(document_slice: dict[str, Any]) -> list[dict[str, Any]]:
    from app.services.period_service import parse_direct_month_label, parse_period

    out: list[dict[str, Any]] = []
    for row in document_slice.get("monthly") or []:
        period_raw = parse_direct_month_label(str(row.get("month") or ""))
        try:
            period_display = parse_period(period_raw)["display"]
        except ValueError:
            period_display = period_raw
        payload: dict[str, Any] = {
            "period": period_display,
            "budget": row.get("cost"),
            "clicks": row.get("clicks"),
            "leads": row.get("leads"),
            "leads_forms": row.get("forms"),
            "leads_messenger": row.get("messengers"),
        }
        if row.get("impressions"):
            payload["impressions"] = row.get("impressions")
        out.append(payload)
    return out


def build_slice_markdown(document_slice: dict[str, Any]) -> str:
    """Human-readable compact summary for AI and document preview."""
    lines: list[str] = ["# Срез данных: Яндекс Директ", ""]
    if document_slice.get("client_label"):
        lines.append(f"**Клиент:** {document_slice['client_label']}")
    if document_slice.get("report_title"):
        lines.append(f"**Отчёт:** {document_slice['report_title']}")
    if document_slice.get("period"):
        lines.append(f"**Период:** {document_slice['period']}")
    goals = document_slice.get("goals") or {}
    if goals:
        lines.append(
            f"**Цели:** лиды = форма + мессенджер ({goals.get('forms', 'форма')} + {goals.get('messengers', 'мессенджер')})."
        )
    lines.append("")

    monthly = document_slice.get("monthly") or []
    if monthly:
        lines.append("## Динамика по месяцам")
        lines.append("| Месяц | Расход | Клики | Лиды | Форма | Мессенджер | CPL |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
        for row in monthly:
            lines.append(
                "| {month} | {cost} | {clicks} | {leads} | {forms} | {msg} | {cpl} |".format(
                    month=row.get("month", "—"),
                    cost=_fmt_money(row.get("cost")),
                    clicks=_fmt_int(row.get("clicks")),
                    leads=_fmt_int(row.get("leads")),
                    forms=_fmt_int(row.get("forms")),
                    msg=_fmt_int(row.get("messengers")),
                    cpl=_fmt_money(row.get("cpl")),
                )
            )
        lines.append("")

    totals = document_slice.get("totals") or {}
    if totals:
        lines.extend(
            [
                "## Итого за период",
                f"- Расход: {_fmt_money(totals.get('cost'))}",
                f"- Клики: {_fmt_int(totals.get('clicks'))}",
                f"- Лиды (всего): {_fmt_int(totals.get('leads') or totals.get('conversions'))}",
            ]
        )
        if totals.get("forms") is not None:
            lines.append(f"- Форма: {_fmt_int(totals.get('forms'))}")
        if totals.get("messengers") is not None:
            lines.append(f"- Мессенджер: {_fmt_int(totals.get('messengers'))}")
        if totals.get("cpl") is not None:
            lines.append(f"- CPL: {_fmt_money(totals.get('cpl'))}")
        elif totals.get("cpa") is not None:
            lines.append(f"- CPA: {_fmt_money(totals.get('cpa'))}")
        lines.append("")

    campaigns = document_slice.get("campaigns") or []
    if campaigns:
        ranked = sorted(
            campaigns,
            key=lambda c: (c.get("cost") is None, -(c.get("cost") or 0)),
        )
        lines.append("## Кампании (агрегат за период)")
        if document_slice.get("report_type") == "conditions_by_month":
            lines.append("| Кампания | Расход | Клики | Лиды | Форма | Мессенджер | CPL |")
            lines.append("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
            for row in ranked[:20]:
                lines.append(
                    "| {name} | {cost} | {clicks} | {leads} | {forms} | {msg} | {cpl} |".format(
                        name=(row.get("campaign_name") or "—")[:50],
                        cost=_fmt_money(row.get("cost")),
                        clicks=_fmt_int(row.get("clicks")),
                        leads=_fmt_int(row.get("leads") or row.get("conversions")),
                        forms=_fmt_int(row.get("forms")),
                        msg=_fmt_int(row.get("messengers")),
                        cpl=_fmt_money(row.get("cpl") or row.get("cpa")),
                    )
                )
        else:
            lines.append("| Кампания | Расход | Клики | Конверсии | CPA |")
            lines.append("| --- | ---: | ---: | ---: | ---: |")
            for row in ranked[:30]:
                lines.append(
                    "| {name} | {cost} | {clicks} | {conv} | {cpa} |".format(
                        name=(row.get("campaign_name") or "—")[:60],
                        cost=_fmt_money(row.get("cost")),
                        clicks=_fmt_int(row.get("clicks")),
                        conv=_fmt_int(row.get("conversions")),
                        cpa=_fmt_money(row.get("cpa")),
                    )
                )
        if len(campaigns) > 20:
            lines.append(f"\n_В файле агрегировано {len(campaigns)} кампаний._")
        lines.append("")

    cond_summary = document_slice.get("conditions_summary") or {}
    top_cond = cond_summary.get("top_by_spend") or []
    if top_cond:
        lines.append("## Условия показа (топ по расходу)")
        lines.append("| Условие | Кампания | Расход | Лиды | CPL |")
        lines.append("| --- | --- | ---: | ---: | ---: |")
        for row in top_cond[:15]:
            lines.append(
                "| {cond} | {camp} | {cost} | {leads} | {cpl} |".format(
                    cond=(row.get("condition") or "—")[:55],
                    camp=(row.get("campaign_name") or "—")[:35],
                    cost=_fmt_money(row.get("cost")),
                    leads=_fmt_int(row.get("leads")),
                    cpl=_fmt_money(row.get("cpl")),
                )
            )
        unique = cond_summary.get("unique_conditions")
        if unique:
            lines.append(f"\n_Уникальных условий в файле: {unique}._")
        lines.append("")

    insights = document_slice.get("insights") or build_slice_insights(document_slice)
    cond_insights = cond_summary.get("insights") or build_conditions_insights(cond_summary)
    if cond_insights:
        for item in cond_insights:
            insights = list(insights) if insights else []
            if not any(i.get("id") == item.get("id") for i in insights):
                insights.append(item)
    if insights:
        lines.append("## Наблюдения (для проверки)")
        for item in insights:
            lines.append(f"- **{item.get('title')}:** {item.get('detail')}")
        lines.append("")

    sheets = document_slice.get("sheets_parsed") or []
    if sheets:
        lines.append(f"_Разобраны листы: {', '.join(sheets)}._")
    warnings = document_slice.get("warnings") or []
    if warnings:
        lines.append("")
        for w in warnings:
            lines.append(f"- ⚠ {w}")
    return "\n".join(lines).strip()


def slice_to_metrics_payload(document_slice: dict[str, Any]) -> dict[str, Any]:
    """Single-period KPI from totals (full-range import uses slice_monthly_to_metrics_list)."""
    out: dict[str, Any] = {}
    period = document_slice.get("period")
    if period:
        out["period"] = str(period).replace("-", "—")
    totals = document_slice.get("totals") or {}
    if totals.get("cost") is not None:
        out["budget"] = totals["cost"]
    if totals.get("clicks") is not None:
        out["clicks"] = int(totals["clicks"])
    leads = totals.get("leads")
    if leads is None:
        leads = totals.get("conversions")
    if leads is not None:
        out["leads"] = int(leads)
    if totals.get("forms") is not None:
        out["leads_forms"] = int(totals["forms"])
    if totals.get("messengers") is not None:
        out["leads_messenger"] = int(totals["messengers"])
    return out


def try_parse_yandex_direct_xlsx(path: Path) -> dict[str, Any] | None:
    """
    If path looks like a Yandex Direct statistics export, return
    {document_slice, extracted_text, sheets_parsed}.
    Otherwise None (caller should use flat xlsx dump).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    best: dict[str, Any] | None = None
    best_campaigns = 0
    meta_global: dict[str, Any] = {}
    sheets_parsed: list[str] = []
    warnings: list[str] = []

    try:
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            meta_global.update({k: v for k, v in _scan_meta(rows).items() if v})

            conditions_parsed: dict[str, Any] | None = None
            for i, row in enumerate(rows[:15]):
                if not _is_conditions_report(row):
                    continue
                col_map = _map_conditions_columns(row)
                conditions_parsed = _parse_conditions_sheet(rows, header_row=i, col_map=col_map)
                if conditions_parsed:
                    break

            if conditions_parsed:
                best = {
                    **meta_global,
                    **conditions_parsed,
                    "format": "yandex_direct_xlsx",
                    "source_sheet": sheet.title,
                }
                sheets_parsed = [sheet.title]
                break

            parsed = _parse_sheet_table(rows)
            if not parsed:
                continue
            n_campaigns = len(parsed.get("campaigns") or [])
            if n_campaigns >= best_campaigns or (
                n_campaigns == best_campaigns and parsed.get("totals") and not (best or {}).get("totals")
            ):
                best_campaigns = n_campaigns
                best = {
                    **meta_global,
                    "format": "yandex_direct_xlsx",
                    "report_type": "campaign_summary",
                    "source_sheet": sheet.title,
                    "totals": parsed.get("totals"),
                    "monthly": parsed.get("monthly") or [],
                    "campaigns": parsed.get("campaigns"),
                    "columns_detected": parsed.get("columns"),
                }
                sheets_parsed = [sheet.title]
            elif parsed.get("totals") and best and not best.get("totals"):
                best["totals"] = parsed["totals"]
                sheets_parsed.append(sheet.title)
    finally:
        wb.close()

    if not best or (not best.get("campaigns") and not best.get("totals") and not best.get("monthly")):
        return None

    if not best.get("period"):
        best["period"] = _period_from_filename(path)
    if not best.get("period"):
        warnings.append(
            "Период не найден в файле — укажите период вручную в «Ручных метриках»."
        )
    if not best.get("totals"):
        warnings.append(
            "Строка «Итого» не найдена — KPI могут быть неполными; проверьте цифры."
        )

    best["sheets_parsed"] = sheets_parsed
    best["warnings"] = warnings
    best["insights"] = build_slice_insights(best)
    markdown = build_slice_markdown(best)
    return {
        "document_slice": best,
        "extracted_text": markdown,
        "sheets_parsed": sheets_parsed,
    }
