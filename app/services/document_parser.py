"""Extract text from uploaded documents for audit materials."""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

SUPPORTED_DOCUMENT_EXTENSIONS = {".txt", ".md", ".csv", ".xlsx", ".pdf", ".docx"}

_SEMANTICS_SHEET_RE = re.compile(
    r"текст|фраз|минус|регион|словар|keyword|ключ|объявлен|названи|слово",
    re.IGNORECASE,
)
_KPI_COLUMN_RE = re.compile(r"расход|клик|конверс|итого|заявк|лид", re.IGNORECASE)


def parse_document(path: Path, *, original_name: str | None = None) -> dict[str, Any]:
    """Parse document and return extracted_text + metadata."""
    suffix = path.suffix.lower()
    name = original_name or path.name

    if suffix not in SUPPORTED_DOCUMENT_EXTENSIONS:
        raise ValueError(f"Неподдерживаемый формат файла: {suffix or 'без расширения'}")

    if suffix in {".txt", ".md"}:
        text = path.read_text(encoding="utf-8", errors="replace").strip()
        return {"extracted_text": text, "parser": suffix.lstrip(".")}

    if suffix == ".csv":
        return _parse_csv(path)

    if suffix == ".xlsx":
        return _parse_xlsx(path)

    if suffix == ".pdf":
        return _parse_pdf(path)

    if suffix == ".docx":
        return _parse_docx(path)

    raise ValueError(f"Неподдерживаемый формат: {suffix}")


def _parse_csv(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(raw))
    rows = [" | ".join(cell.strip() for cell in row) for row in reader if any(cell.strip() for cell in row)]
    text = "\n".join(rows).strip()
    return {"extracted_text": text, "parser": "csv", "rows": len(rows)}


def _parse_xlsx(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Для .xlsx нужен пакет openpyxl") from exc

    from app.services.yandex_direct_xlsx_service import try_parse_yandex_direct_xlsx

    direct = try_parse_yandex_direct_xlsx(path)
    if direct:
        slice_data = direct.get("document_slice") or {}
        return {
            "extracted_text": direct.get("extracted_text") or "",
            "parser": "yandex_direct_xlsx",
            "document_slice": slice_data,
            "sheets": len(slice_data.get("sheets_parsed") or []),
            "format": slice_data.get("format"),
        }

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names_list = list(wb.sheetnames or [])
    sheet_count = len(sheet_names_list)
    if _workbook_looks_like_semantics_export(wb):
        wb.close()
        wb = load_workbook(path, data_only=True)
        semantics = _parse_xlsx_semantics_export(wb)
        wb.close()
        if semantics:
            return semantics
        wb = load_workbook(path, read_only=True, data_only=True)

    is_semantics_like = _workbook_looks_like_semantics_export(wb)
    chunks: list[str] = [
        "# Документ Excel",
        "",
    ]
    if is_semantics_like:
        chunks.extend([
            "Справочник фраз / минус-слов / текстов (не мастер-отчёт с KPI). "
            "Данные ниже можно использовать для семантики и рекомендаций по минус-словам.",
            "",
        ])
    else:
        chunks.extend([
            "Не удалось распознать формат «Мастер отчётов» Яндекс Директа "
            "(нужны колонки Расход, Клики, Конверсии и строка «Итого»). "
            "Ниже — сырой текст листов; для KPI загрузите отчёт по кампаниям на вкладку «Директ».",
            "",
        ])
    max_chars = 120_000
    total_chars = 0
    for sheet in wb.worksheets:
        chunks.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if not cells:
                continue
            line = " | ".join(cells)
            chunks.append(line)
            total_chars += len(line)
            if total_chars >= max_chars:
                chunks.append("… (файл обрезан)")
                wb.close()
                payload = {
                    "extracted_text": "\n".join(chunks).strip(),
                    "parser": "xlsx",
                    "sheets": sheet_count,
                    "truncated": True,
                }
                if is_semantics_like:
                    payload["document_kind"] = "direct_semantics_export"
                    payload["document_slice"] = {
                        "format": "direct_semantics_export",
                        "sheets": sheet_names_list,
                    }
                return payload
    wb.close()
    text = "\n".join(chunks).strip()
    payload: dict[str, Any] = {"extracted_text": text, "parser": "xlsx", "sheets": sheet_count}
    if is_semantics_like:
        payload["document_kind"] = "direct_semantics_export"
        payload["document_slice"] = {
            "format": "direct_semantics_export",
            "sheets": sheet_names_list,
        }
    return payload


def _workbook_looks_like_semantics_export(wb) -> bool:
    names = list(wb.sheetnames or [])
    if not names:
        return False
    name_hits = sum(1 for n in names if _SEMANTICS_SHEET_RE.search(str(n)))
    if name_hits >= 1 and name_hits >= max(1, len(names) // 3):
        return True
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(max_row=8, values_only=True):
            header = " ".join(str(c) for c in row if c is not None)
            if _KPI_COLUMN_RE.search(header):
                return False
            if _SEMANTICS_SHEET_RE.search(header):
                return True
    return False


def _parse_xlsx_semantics_export(wb) -> dict[str, Any] | None:
    if not _workbook_looks_like_semantics_export(wb):
        return None
    sheet_names = list(wb.sheetnames or [])
    chunks: list[str] = [
        "# Справочник фраз, минус-слов и текстов (Excel Директа)",
        "",
        "Тип: не мастер-отчёт с KPI. Используй для зоны «Семантика» — фразы, минус-слова, "
        "тексты объявлений, регионы. Цифры расхода/лидов — только из мастер-отчёта на «Директ».",
        "",
    ]
    max_chars = 120_000
    total_chars = 0
    for sheet in wb.worksheets:
        chunks.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if not cells:
                continue
            line = " | ".join(cells)
            chunks.append(line)
            total_chars += len(line)
            if total_chars >= max_chars:
                chunks.append("… (файл обрезан)")
                return {
                    "extracted_text": "\n".join(chunks).strip(),
                    "parser": "yandex_direct_semantics_xlsx",
                    "document_kind": "direct_semantics_export",
                    "document_slice": {
                        "format": "direct_semantics_export",
                        "sheets": sheet_names,
                    },
                    "sheets": len(sheet_names),
                    "truncated": True,
                }
    return {
        "extracted_text": "\n".join(chunks).strip(),
        "parser": "yandex_direct_semantics_xlsx",
        "document_kind": "direct_semantics_export",
        "document_slice": {
            "format": "direct_semantics_export",
            "sheets": sheet_names,
        },
        "sheets": len(sheet_names),
    }


def _parse_pdf(path: Path) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError("Для .pdf нужен пакет pypdf") from exc

    reader = PdfReader(str(path))
    pages = []
    for page in reader.pages:
        pages.append((page.extract_text() or "").strip())
    text = "\n\n".join(p for p in pages if p).strip()
    return {"extracted_text": text, "parser": "pdf", "pages": len(reader.pages)}


def _parse_docx(path: Path) -> dict[str, Any]:
    try:
        from docx import Document
    except ImportError as exc:
        raise ValueError("Для .docx нужен пакет python-docx") from exc

    doc = Document(str(path))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    text = "\n".join(paragraphs).strip()
    return {"extracted_text": text, "parser": "docx", "paragraphs": len(paragraphs)}
