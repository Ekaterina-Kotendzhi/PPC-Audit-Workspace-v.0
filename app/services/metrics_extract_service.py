"""Extract KPI fields from text notes, documents, and xlsx uploads."""
from __future__ import annotations

import re
from datetime import datetime, timezone
from typing import Any

from app.models import AuditProject
from app.services.file_service import file_url_to_path

METRICS_EXTRACT_LABELS = {
    "period": "Период",
    "budget": "Бюджет",
    "clicks": "Клики",
    "leads": "Заявки",
    "sales": "Продажи",
    "revenue": "Выручка",
}

KPI_FIELDS = ("period", "budget", "clicks", "leads", "sales", "revenue")


def parse_num(raw: str) -> float | None:
    value = (raw or "").strip().replace(" ", "").replace(",", ".")
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def labeled_metric_lines(lower: str) -> dict[str, Any]:
    """Parse «Бюджет: 164 200», «Заявки: 132» lines."""
    out: dict[str, Any] = {}
    budgets: list[float] = []
    clicks: list[int] = []
    leads: list[int] = []
    sales: list[int] = []
    revenues: list[float] = []
    for line in lower.splitlines():
        line_s = line.strip()
        if m := re.match(r"^период\s*:\s*(.+)$", line_s):
            out["period"] = m.group(1).strip().replace("-", "—")
        if m := re.match(r"^бюджет\s*:\s*([\d\s.,]+)", line_s):
            val = parse_num(m.group(1))
            if val is not None:
                budgets.append(val)
        if m := re.match(r"^расход\s*:\s*([\d\s.,]+)", line_s):
            val = parse_num(m.group(1))
            if val is not None:
                budgets.append(val)
        if m := re.match(r"^клики\s*:\s*([\d\s.,]+)", line_s):
            val = parse_num(m.group(1))
            if val is not None:
                clicks.append(int(val))
        if m := re.match(r"^заявки\s*:\s*([\d\s.,]+)", line_s):
            val = parse_num(m.group(1))
            if val is not None:
                leads.append(int(val))
        if m := re.match(r"^продажи\s*:\s*([\d\s.,]+)", line_s):
            val = parse_num(m.group(1))
            if val is not None:
                sales.append(int(val))
        if m := re.match(r"^выручка\s*:\s*([\d\s.,]+)", line_s):
            val = parse_num(m.group(1))
            if val is not None:
                revenues.append(val)
    if budgets:
        out["budget"] = max(budgets)
    if clicks:
        out["clicks"] = max(clicks)
    if leads:
        out["leads"] = max(leads)
    if sales:
        out["sales"] = max(sales)
    if revenues:
        out["revenue"] = max(revenues)
    return out


def extract_metrics_payload_from_text(text: str) -> dict[str, Any]:
    lower = (text or "").lower()
    out: dict[str, Any] = dict(labeled_metric_lines(lower))

    period_match = re.findall(r"\b\d{2}\.\d{2}\.\d{4}\s*[-—]\s*\d{2}\.\d{2}\.\d{4}\b", lower)
    if period_match:
        out["period"] = period_match[-1].replace("-", "—")
    else:
        month_match = re.findall(
            r"\b(январ[ьяе]|феврал[ьяе]|март[ае]?|апрел[ьяе]|ма[йяе]|июн[ьяе]|июл[ьяе]|август[ае]?|сентябр[ьяе]|октябр[ьяе]|ноябр[ьяе]|декабр[ьяе])\s+\d{4}\b",
            lower,
        )
        if month_match:
            out["period"] = month_match[-1]

    total_line = next((line for line in lower.splitlines() if "итого" in line), "")
    if total_line:
        tokens = re.split(r"[^\d,.\s]+", total_line)
        candidates: list[str] = []
        for token in tokens:
            cleaned = token.strip()
            if not cleaned:
                continue
            for part in cleaned.split():
                part = part.strip()
                if re.fullmatch(r"\d+(?:[.,]\d+)?", part):
                    candidates.append(part)
        parsed = [parse_num(n) for n in candidates]
        parsed = [n for n in parsed if n is not None]
        if len(parsed) >= 1:
            out["budget"] = parsed[0]
        if len(parsed) >= 2:
            out["clicks"] = int(parsed[1])
        if len(parsed) >= 3:
            out["leads"] = int(parsed[2])

    if "budget" not in out:
        m = re.search(r"(расход|бюджет)[^\d]{0,20}(\d[\d\s]*[.,]?\d*)", lower)
        if m:
            val = parse_num(m.group(2))
            if val is not None:
                out["budget"] = val
    if "clicks" not in out:
        m = re.search(r"клик[аиы]?[^\d]{0,20}(\d[\d\s]*[.,]?\d*)", lower)
        if m:
            val = parse_num(m.group(1))
            if val is not None:
                out["clicks"] = int(val)
    if "leads" not in out:
        lead_candidates: list[float] = []
        for m in re.finditer(r"\bзаявки\s*:\s*(\d[\d\s]*[.,]?\d*)", lower):
            val = parse_num(m.group(1))
            if val is not None:
                lead_candidates.append(val)
        for m in re.finditer(r"(?:^|\n)[^\n]{0,80}?\bзаявки\b[^\d]{0,8}(\d[\d\s]*[.,]?\d*)", lower):
            val = parse_num(m.group(1))
            if val is not None and val >= 3:
                lead_candidates.append(val)
        if lead_candidates:
            out["leads"] = int(max(lead_candidates))

    return out


def _skip_xlsx_kpi_rescan(material: Any) -> bool:
    """Non–Master Report uploads should not re-parse large xlsx on every audit GET."""
    title = (getattr(material, "title", None) or "").lower()
    if any(k in title for k in ("поисков", "запрос", "семантик", "keyword", "минус")):
        return True
    return False


def extract_metrics_from_xlsx(path: Any, *, allow_generic_scan: bool = True) -> dict[str, Any]:
    from pathlib import Path

    from app.services.yandex_direct_xlsx_service import (
        slice_to_metrics_payload,
        try_parse_yandex_direct_xlsx,
    )

    p = Path(path) if not isinstance(path, Path) else path
    direct = try_parse_yandex_direct_xlsx(p)
    if direct and direct.get("document_slice"):
        payload = slice_to_metrics_payload(direct["document_slice"])
        if payload:
            return payload

    if not allow_generic_scan:
        return {}

    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    payload: dict[str, Any] = {}
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception:
        return {}

    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [("" if v is None else str(v)).strip() for v in row]
                line = " ".join([c for c in cells if c])
                if not line:
                    continue
                parsed = extract_metrics_payload_from_text(line)
                for key, value in parsed.items():
                    if key not in payload and value not in (None, ""):
                        payload[key] = value

                low_cells = [c.lower() for c in cells]
                if any("итого" in c for c in low_cells):
                    nums = []
                    for c in cells:
                        parts = re.findall(r"\d+(?:[.,]\d+)?", c.replace(" ", ""))
                        nums.extend(parts)
                    parsed_nums = [parse_num(n) for n in nums]
                    parsed_nums = [n for n in parsed_nums if n is not None]
                    if len(parsed_nums) >= 1 and "budget" not in payload:
                        payload["budget"] = parsed_nums[0]
                    if len(parsed_nums) >= 2 and "clicks" not in payload:
                        payload["clicks"] = int(parsed_nums[1])
                    if len(parsed_nums) >= 3 and "leads" not in payload:
                        payload["leads"] = int(parsed_nums[2])
    finally:
        wb.close()

    return payload


def metrics_extract_preview_lines(payload: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for key in KPI_FIELDS:
        if key not in payload or payload[key] in (None, ""):
            continue
        label = METRICS_EXTRACT_LABELS.get(key, key)
        lines.append(f"{label}: {payload[key]}")
    return lines


def _normalize_ts(value: datetime | None) -> datetime:
    if value is None:
        return datetime.min.replace(tzinfo=timezone.utc)
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value


def _sorted_materials(project: AuditProject, types: set[str]):
    return sorted(
        [
            m
            for m in (project.materials or [])
            if m.type in types and not bool(getattr(m, "excluded_from_analysis", False))
        ],
        key=lambda item: _normalize_ts(getattr(item, "updated_at", None) or getattr(item, "created_at", None)),
        reverse=True,
    )


def collect_metrics_from_material_types(project: AuditProject, types: set[str]) -> dict[str, Any]:
    """Newest material wins per field within the given types."""
    payload: dict[str, Any] = {}
    for material in _sorted_materials(project, types):
        text = (material.extracted_text or material.raw_content or "").strip()
        extracted = extract_metrics_payload_from_text(text)

        if material.type == "document":
            from app.services.material_helpers import document_slice_from_material
            from app.services.yandex_direct_xlsx_service import slice_to_metrics_payload

            slice_data = document_slice_from_material(material)
            if slice_data and slice_data.get("format") == "yandex_direct_xlsx":
                for key, value in slice_to_metrics_payload(slice_data).items():
                    if key not in extracted and value not in (None, ""):
                        extracted[key] = value
            elif getattr(material, "file_url", None) and not _skip_xlsx_kpi_rescan(material):
                local_path = file_url_to_path(material.file_url)
                if local_path and str(local_path).lower().endswith(".xlsx"):
                    xlsx_extracted = extract_metrics_from_xlsx(local_path, allow_generic_scan=False)
                    for key, value in xlsx_extracted.items():
                        if key not in extracted and value not in (None, ""):
                            extracted[key] = value

        for key, value in extracted.items():
            if key not in payload and value not in (None, ""):
                payload[key] = value

        title_low = (material.title or "").lower()
        simple_num = parse_num(text)
        if simple_num is not None:
            if "budget" not in payload and "бюдж" in title_low:
                payload["budget"] = simple_num
            if "clicks" not in payload and "клик" in title_low:
                payload["clicks"] = int(simple_num)
            if "leads" not in payload and ("заяв" in title_low or "конверс" in title_low):
                payload["leads"] = int(simple_num)
    return payload


def collect_metrics_extract_payload(
    project: AuditProject,
    *,
    note_id: int | None = None,
    include_documents: bool = True,
) -> dict[str, Any]:
    """Notes first (marketer intent), then documents fill gaps. With note_id — only that note."""
    if note_id is not None:
        note = next(
            (m for m in (project.materials or []) if m.id == note_id and m.type == "text_note"),
            None,
        )
        if not note:
            return {}
        text = (note.extracted_text or note.raw_content or "").strip()
        return extract_metrics_payload_from_text(text)

    payload = collect_metrics_from_material_types(project, {"text_note"})
    if include_documents:
        for key, value in collect_metrics_from_material_types(project, {"document"}).items():
            if key not in payload and value not in (None, ""):
                payload[key] = value
    return payload
